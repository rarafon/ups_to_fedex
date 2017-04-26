import openpyxl, ffile

import re
import datetime

rates = None

def save_fedex_rates(rates):
	ffile.save_fedex_rates(rates)
	return rates

def process_excel_fedex(file_name="fedex_rates.xlsx"):
	ffile.move_dir("data")

	wb = openpyxl.load_workbook(file_name)
	sheetnames_list = wb.get_sheet_names()

	rate_dic = {}

	for delivery_name, para_list in fedex_rate_proc_index.items():
		sheet = wb.get_sheet_by_name(delivery_name)
		rates = proc_sheet_for_rates(sheet, *para_list)
		# print(delivery_name)
		# print(rates)
		rate_dic[delivery_name] = rates

	# sheet = wb.get_sheet_by_name('Standard Overnight')
	# rates = proc_sheet_for_rates(sheet)
	# print(rates)

	ffile.dir_back()

	return rate_dic

def open_rates():
	global rates
	r = ffile.open_fedex_rates()
	rates = r
	return r

column_letters = [
	'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 
	'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z'
]

def get_zone_dic(zone_list, col_letters_list, zone_limit = 8):
	zone_dic = {}

	for i, letter in enumerate(col_letters_list):
		zone_dic[letter] = zone_list[i]

		if zone_list[i] == zone_limit:
			break
	
	return zone_dic

def proc_sheet_for_rates(sheet, zone_row_num=2, rate_start_row_num=6, total_columns=11):
# Skip row 3, b/c Envelopes are usually not used for overnight.
	max_rows = sheet.max_row
	max_columns = sheet.max_column
	col_letters = column_letters[:total_columns]

	zone_list = []
	# Get Zones on second row
	for letter in col_letters:
		zone = sheet[letter + str(zone_row_num)].value
		zone_list.append(zone)

		# if zone == 8:
			# break
	zone_dic = get_zone_dic(zone_list, col_letters)

	##Start translate rows and weight for rate
	weight_dic = {}
	for row in range(rate_start_row_num, max_rows + 1):
		weight = 1
		for letter, zone in zone_dic.items():
			if letter == 'A':
				weight = sheet[letter + str(row)].value
				weight_dic[weight] = {}
			else:
				price = sheet[letter + str(row)].value
				weight_dic[weight][zone] = price
	return weight_dic

fedex_rate_proc_index = {	
# This is for the parameters to use in proc_sheet_for_rates
# Will be iterated to match with the sheet names of the
# 	fedex excel files of the rates in each sheet.
	'Priority Overnight': [2,6,10],
	'Standard Overnight': [2,6,10],
	'2 Day AM': [2,6,10],
	'2 Day': [2,6,11],
	'Express Saver (3 Day)': [2,3,8],
	'Ground': [2,3,8],
	'Home Delivery': [2,3,8],
	'Smart Post 1-16 oz': [2,3,8],
	'Smart Post 1-70 lbs': [2,3,8],

	#Tese are delivery to weird areas

	# 'Deferred Smart Post 1-16 oz': None,
	# 'Deferred Smart Post 1-70 lbs': None,
}

def get_rate(fedex_service_name, weight, zone):
	weight = int(weight)
	zone = int(zone)
	global rates
	if rates == None:
		rates = process_excel_fedex()
		save_fedex_rates(rates)
		rates = open_rates()
	return rates[fedex_service_name][weight][zone]


def calc_ground_commercial(weight, zone):
	return get_rate('Ground', weight, zone)

def calc_ground_residential(weight, zone):
	return calc_ground_commercial(weight, zone) + calc_residential_charge(weight, zone)

def calc_2_day_air_commercial(weight, zone):
	return get_rate('2 Day', weight, zone)

def calc_smart_post_1lb_plus(weight, zone):
	return get_rate('Smart Post 1-70 lbs', weight, zone)

def calc_residential_charge(weight, zone):
	res_surcharge = 3.45
	discount = 0.50
	return res_surcharge - discount

def calc_oversize_charge(weight, zone):
	oversize_charge = 72.50
	discount = 0.00
	return oversize_charge - discount

def calc_add_handling(weight, zone):
	add_handling = 11.00
	discount = add_handling * 0.25
	return add_handling - discount

def calc_delivery_area_surcharge(service_type, residential, extended):
	# type refers to residential or commercial
	delivery_area_surcharge = 0
	if service_type == "Ground":
		if residential:
			if extended:
				delivery_area_surcharge = 4.2
			else:
				delivery_area_surcharge = 3.9
		else:
			delivery_area_surcharge = 2.45
	elif service_type == 'Priority Overnight' or service_type =='Standard Overnight' or service_type == '2 Day AM' or service_type == '2 Day':
		if residential:
			if extended:
				delivery_area_surcharge = 4.2
			else:
				delivery_area_surcharge = 3.9
		else:
			delivery_area_surcharge = 2.6
	elif service_type == "Home Delivery":
		if extended:
			delivery_area_surcharge = 4.2
		else:
			delivery_area_surcharge = 3.35
	elif service_type == 'Smart Post 1-16 oz' or service_type == 'Smart Post 1-70 lbs':
		if extended:
			delivery_area_surcharge = 1.50
		else:
			delivery_area_surcharge = 1.00
	else:
		msg = "Unknown service_type " + service_type
		print(msg)

	discount = delivery_area_surcharge * 0.25

	return delivery_area_surcharge - discount

def calc_signature(weight, zone):
	signature_rate = 4.5
	discount = signature_rate * 0.25
	return signature_rate - discount

def calc_nonmachinable_charge(weight, zone):
	nonmachinable_charge = 2.5
	return nonmachinable_charge

def calc_fuel_surcharge(self, date, charge_type_list):
	print(date)
	print(charge_type_list)

# Fuel rate percentages are stored in [Express_value, Ground_value]

fuel_rate_index = {
# Each date is for the entire week.
	"1/2/2017": [2.50, 4.00],
	"1/9/2017": [2.50, 4.00],
	"1/16/2017": [2.50, 4.00],
	"1/23/2017": [2.50, 4.00],
	"1/30/2017": [2.50, 4.00],
	"2/6/2017": [3.50, 4.50],
	"2/13/2017": [3.50, 4.25], 
	"2/20/2017": [3.50, 4.50],
	"2/27/2017": [3.75, 4.50],
	"3/6/2017": [3.75, 4.50],
	"3/13/2017": [3.50, 4.50],
	"3/20/2017": [3.25, 4.50],
	"3/27/2017": [2.75, 4.25], 
	"4/3/2017": [2.75, 4.25],
	"4/10/2017": [3.00, 4.25],
	"4/17/2017": [3.50, 4.50],
	"4/24/2017": [3.75, 4.50],
}

def get_date_from_string(date_string):
	#Format will be in mm/dd/yyyy format
	date_re_pattern = "(?P<month>\d+)\\(?P<day>\d+)\\(?P<year>\d+)"
	r = re.match(r"(?P<month>\d+)\/(?P<day>\d+)\/(?P<year>\d+)", date_string)
	month = int(r.group('month'))
	year = int(r.group('year'))
	day = int(r.group('day'))
	return {"year": year, "month": month, "day": day,}

def get_string_from_date(year, month, day):
	return str(month) + "/" + str(day) + "/" + str(year)

def fill_fuel_rates():
	# Fills fuel_rate_dic with the rates in fuel_rate_index

	# it fills a total days of fill_num_days + 1
	fill_num_days = 6
	for date_str, rate_list in fuel_rate_index.items():
		date_dic = get_date_from_string(date_str)
		year = date_dic["year"]
		month = date_dic["month"]
		day = date_dic["day"]

		put_data_into_fuel_rate_dic(year, month, day, rate_list)

		d = datetime.date(year, month, day)
		print(d)

		for i in range(1, fill_num_days + 1):
			new_d = d + datetime.timedelta(days=i)
			new_year = new_d.year
			new_month = new_d.month
			new_day = new_d.day
			put_data_into_fuel_rate_dic(new_year, new_month, new_day, rate_list)

def put_data_into_fuel_rate_dic(year, month, day, data):
	if year not in fuel_rate_dic:
		fuel_rate_dic[year] = {}
	if month not in fuel_rate_dic[year]:
		fuel_rate_dic[year][month]= {}
	fuel_rate_dic[year][month][day] = data

def get_fuel_rate(year, month, day, delivery_type):
	fuel_list = fuel_rate_dic[year][month][day]
	if delivery_type == "Express":
		return fuel_list[0]
	elif delivery_type == "Ground":
		return fuel_list[1]
	else:
		msg = "ERROR: delivery type " + delivery_type + " is seen." 
		raise Exception(msg)

fuel_rate_dic = {}

